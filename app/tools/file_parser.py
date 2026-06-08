"""입력 파일 파싱 — PDF / DOCX / XLSX / MD / TXT (Stage 1)."""
from pathlib import Path


def parse(path: str | Path) -> str:
    """파일을 읽어 평문 텍스트로 반환."""
    p = Path(path)
    suffix = p.suffix.lower()

    if suffix == ".pdf":
        return _parse_pdf(p)
    if suffix in (".docx", ".doc"):
        return _parse_docx(p)
    if suffix in (".xlsx", ".xls"):
        return _parse_xlsx(p)
    if suffix in (".md", ".txt", ".csv"):
        return p.read_text(encoding="utf-8", errors="replace")

    raise ValueError(f"지원하지 않는 파일 형식: {suffix}")


def _parse_pdf(path: Path) -> str:
    import fitz  # PyMuPDF
    doc = fitz.open(str(path))
    return "\n".join(page.get_text() for page in doc)


def _parse_docx(path: Path) -> str:
    from docx import Document
    doc = Document(str(path))
    return "\n".join(p.text for p in doc.paragraphs if p.text.strip())


def _parse_xlsx(path: Path) -> str:
    from openpyxl import load_workbook
    wb = load_workbook(str(path), read_only=True, data_only=True)
    lines = []
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            row_text = "\t".join(str(c) if c is not None else "" for c in row)
            if row_text.strip():
                lines.append(row_text)
    return "\n".join(lines)
