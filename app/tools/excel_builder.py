"""TC Excel 산출 빌더 (Stage 7, doc/03-tc-schema.md §6)."""
from __future__ import annotations
import re
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# openpyxl이 거부하는 제어문자 (0x00-08, 0x0b, 0x0c, 0x0e-1f) 제거용
# injection/경계 시험 TC 본문에 제어문자·널바이트가 섞여 IllegalCharacterError 발생 → 정화
_ILLEGAL_XLSX_RE = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")


def _san(val: Any) -> Any:
    """셀에 쓸 값 정화. 문자열이면 Excel 금지 제어문자를 제거, 그 외는 그대로."""
    if isinstance(val, str):
        return _ILLEGAL_XLSX_RE.sub("", val)
    return val

# 컬럼 정의 (doc/03-tc-schema.md §2)
# screenshot_file: Stage 0 스크린샷과 TC 연결 (파일명만 저장, 절대경로 아님)
_SHEET1_COLS = [
    "tc_id", "대분류", "중분류", "소분류", "scenario",
    "precondition", "expected", "actual", "result", "failure_reason",
    "screenshot_file",
]
_META_COLS = [
    "tc_id", "requirement_id", "design_technique", "source_quote",
    "gen_confidence", "exec_confidence",
    "review_status", "reviewer_note", "reviewer_id",
    "screenshot_file",
]

# 기능 목록 컬럼 (Stage 0 DOM 스캔 결과)
_FEATURE_COLS = [
    "category_major", "category_mid", "category_leaf",
    "implicit_spec", "source_element", "confidence", "screenshot_file",
]
_FEATURE_COL_NAMES = {
    "category_major":  "대분류",
    "category_mid":    "중분류",
    "category_leaf":   "기능명 (소분류)",
    "implicit_spec":   "기능 명세",
    "source_element":  "근거 DOM 요소",
    "confidence":      "신뢰도",
    "screenshot_file": "관련 스크린샷",
}
# 컬럼별 권장 너비 (기능목록 시트용)
_FEATURE_COL_WIDTHS = {
    "category_major":  16,
    "category_mid":    18,
    "category_leaf":   22,
    "implicit_spec":   45,
    "source_element":  28,
    "confidence":      12,
    "screenshot_file": 30,
}

_TECHNIQUE_SHORT: dict[str, str] = {
    "happy_path":       "정상",
    "negative_basic":   "오류",
    "negative_deep":    "심층오류",
    "boundary":         "경계",
    "equivalence":      "동등분할",
    "state_transition": "상태전이",
    "cross_feature":    "기능간연계",
}

_CONFIDENCE_FILLS = {
    "high": PatternFill(fill_type="solid", fgColor="C6EFCE"),  # 연두
    "mid":  PatternFill(fill_type="solid", fgColor="FFEB9C"),  # 노랑
    "low":  PatternFill(fill_type="solid", fgColor="FFC7CE"),  # 빨강
}

_RESULT_FILLS = {
    "pass":    PatternFill(fill_type="solid", fgColor="C6EFCE"),
    "fail":    PatternFill(fill_type="solid", fgColor="FFC7CE"),
    "blocked": PatternFill(fill_type="solid", fgColor="FFEB9C"),
}

_THIN   = Side(style="thin")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _header_style(cell, fill_color: str = "4472C4"):
    cell.fill      = PatternFill(fill_type="solid", fgColor=fill_color)
    cell.font      = Font(bold=True, color="FFFFFF", size=10)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = _BORDER


def _write_sheet(ws, columns: list[str], rows: list[dict], confidence_col: str | None = None):
    ws.freeze_panes = "A2"
    for ci, col in enumerate(columns, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        _header_style(cell)
        ws.column_dimensions[get_column_letter(ci)].width = max(12, len(col) + 4)

    for ri, row in enumerate(rows, 2):
        for ci, col in enumerate(columns, 1):
            val  = row.get(col, "")
            cell = ws.cell(row=ri, column=ci, value=_san(val))
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border    = _BORDER

            if col == "result" and isinstance(val, str):
                fill = _RESULT_FILLS.get(val.lower())
                if fill:
                    cell.fill = fill

            if col == confidence_col and isinstance(val, (int, float)):
                if val >= 0.85:
                    cell.fill = _CONFIDENCE_FILLS["high"]
                elif val >= 0.50:
                    cell.fill = _CONFIDENCE_FILLS["mid"]
                else:
                    cell.fill = _CONFIDENCE_FILLS["low"]

    ws.auto_filter.ref = ws.dimensions


# ── TC Excel (Stage 7 최종 산출) ──────────────────────────────────────────────

def build(
    tcs: list[dict],
    output_path: str | Path,
    meta: dict | None = None,
) -> Path:
    """TC 목록을 받아 tc_final.xlsx 를 생성하고 경로를 반환.

    Args:
        meta: meta.json 내용. None 아니면 '제한사항' 시트를 추가해 발주처용
              시험 환경·캐시 사용·분석 누락 정보를 명시 (박정훈 권고).
    """
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "표준 양식"
    _write_sheet(ws1, _SHEET1_COLS, tcs)
    # screenshot_file 열 너비 조정
    sf_col = get_column_letter(_SHEET1_COLS.index("screenshot_file") + 1)
    ws1.column_dimensions[sf_col].width = 30

    ws2 = wb.create_sheet("AWT_Meta")
    _write_sheet(ws2, _META_COLS, tcs, confidence_col="gen_confidence")

    # ── (D65) TC-Leaf 커버리지 매트릭스 시트 ────────────────────────────
    if tcs:
        ws_cov = wb.create_sheet("커버리지")
        _write_coverage_matrix_sheet(ws_cov, tcs)

    # ── (D60+) 제한사항 시트 — 박정훈 시험 인증 권고 ─────────────────────
    if meta is not None:
        ws3 = wb.create_sheet("제한사항")
        _write_limitations_sheet(ws3, tcs, meta)

    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out))
    return out


def _write_coverage_matrix_sheet(ws, tcs: list[dict]) -> None:
    """TC-Leaf 커버리지 매트릭스 (D65).

    소분류별 TC 수·설계기법 분포·실행결과를 한 눈에 보여주고
    TC 수 부족 leaf를 색상으로 경고한다.
    """
    from collections import defaultdict

    # ── leaf별 집계 ────────────────────────────────────────────────────
    leaf_data: dict[str, dict] = {}
    for tc in tcs:
        leaf = tc.get("소분류") or "(소분류 없음)"
        if leaf not in leaf_data:
            leaf_data[leaf] = {
                "requirement_id": tc.get("requirement_id", ""),
                "중분류":         tc.get("중분류", ""),
                "tc_ids":         [],
                "techniques":     defaultdict(int),
                "results":        defaultdict(int),
            }
        d = leaf_data[leaf]
        tc_id = tc.get("tc_id", "")
        if tc_id:
            d["tc_ids"].append(tc_id)
        tech = tc.get("design_technique", "")
        if tech:
            d["techniques"][tech] += 1
        result = tc.get("result", "")
        if result and result not in ("not_executed", ""):
            d["results"][result] += 1

    sorted_leaves = sorted(
        leaf_data.items(),
        key=lambda x: (x[1].get("requirement_id", ""), x[0]),
    )

    # ── 헤더 ──────────────────────────────────────────────────────────
    COL_DEFS = [
        ("Leaf ID",      10),
        ("소분류",        24),
        ("중분류",        18),
        ("TC 수",          8),
        ("설계기법 분포", 32),
        ("실행 결과",     22),
        ("TC ID 목록",    50),
        ("커버리지",      14),
    ]
    for ci, (h, w) in enumerate(COL_DEFS, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        _header_style(cell, fill_color="1E6B3C")
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = "A2"

    FILL_OK      = PatternFill("solid", fgColor="C6EFCE")  # 연두 — 충분 (4개+)
    FILL_CAUTION = PatternFill("solid", fgColor="FFEB9C")  # 노랑 — 적음 (2~3개)
    FILL_WARN    = PatternFill("solid", fgColor="FFC7CE")  # 빨강 — 부족 (0~1개)
    FONT_WARN    = Font(bold=True, color="9B1C1C", size=10)

    # ── 데이터 행 ──────────────────────────────────────────────────────
    for ri, (leaf, d) in enumerate(sorted_leaves, 2):
        tc_count = len(d["tc_ids"])

        tech_str = " / ".join(
            f"{_TECHNIQUE_SHORT.get(k, k)}:{v}"
            for k, v in sorted(d["techniques"].items(), key=lambda x: -x[1])
        ) or "—"

        result_str = " / ".join(
            f"{k}:{v}" for k, v in sorted(d["results"].items())
        ) or "미실행"

        tc_ids_preview = d["tc_ids"][:12]
        tc_str = ", ".join(tc_ids_preview)
        if len(d["tc_ids"]) > 12:
            tc_str += f"  … 외 {len(d['tc_ids']) - 12}개"

        if tc_count == 0:
            coverage_label, fill = "없음", FILL_WARN
        elif tc_count == 1:
            coverage_label, fill = "부족 (1개)", FILL_WARN
        elif tc_count <= 3:
            coverage_label, fill = f"적음 ({tc_count}개)", FILL_CAUTION
        else:
            coverage_label, fill = f"충분 ({tc_count}개)", FILL_OK

        row_vals = [
            d.get("requirement_id", ""),
            leaf,
            d.get("중분류", ""),
            tc_count,
            tech_str,
            result_str,
            tc_str,
            coverage_label,
        ]
        for ci, val in enumerate(row_vals, 1):
            cell = ws.cell(row=ri, column=ci, value=_san(val))
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = _BORDER

        # TC 수 셀 + 커버리지 셀에 색상
        ws.cell(row=ri, column=4).fill = fill
        cov_cell = ws.cell(row=ri, column=8)
        cov_cell.fill = fill
        if tc_count <= 1:
            cov_cell.font = FONT_WARN

    ws.auto_filter.ref = ws.dimensions

    # ── 요약 (하단) ────────────────────────────────────────────────────
    total  = len(sorted_leaves)
    n_ok   = sum(1 for _, d in sorted_leaves if len(d["tc_ids"]) >= 4)
    n_caut = sum(1 for _, d in sorted_leaves if 2 <= len(d["tc_ids"]) <= 3)
    n_warn = total - n_ok - n_caut
    n_tc   = sum(len(d["tc_ids"]) for _, d in sorted_leaves)

    sr = total + 3
    ws.cell(row=sr,   column=1, value="요약").font = Font(bold=True, size=11)
    for offset, (label, f) in enumerate([
        (f"전체 leaf {total}개  /  총 TC {n_tc}개",         None),
        (f"커버리지 충분 (TC 4개 이상): {n_ok}개",          FILL_OK),
        (f"커버리지 적음 (TC 2~3개): {n_caut}개",           FILL_CAUTION),
        (f"커버리지 부족 (TC 0~1개): {n_warn}개 — 보완 필요", FILL_WARN),
    ], 1):
        cell = ws.cell(row=sr + offset, column=2, value=label)
        cell.font = Font(bold=(f is not None), size=10)
        if f:
            cell.fill = f


def _write_limitations_sheet(ws, tcs: list[dict], meta: dict) -> None:
    """발주처 시험 성적서 첨부용 — 시험 환경·캐시·분석 누락 정보를 표 형태로 기록."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    HDR_FILL = PatternFill("solid", fgColor="1E40AF")
    HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
    SUB_FILL = PatternFill("solid", fgColor="F1F5F9")
    SUB_FONT = Font(bold=True, color="1E293B", size=10)
    KEY_FONT = Font(bold=True, color="475569", size=10)
    THIN = Side(border_style="thin", color="E2E8F0")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    row = 1

    # ── 제목 ────────────────────────────────────────────────────────────
    ws.cell(row=row, column=1, value="시험 제한사항 및 환경 정보").font = Font(
        bold=True, size=14, color="1E293B"
    )
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    row += 1
    ws.cell(row=row, column=1,
            value="(이 문서는 자동 생성된 시험 결과의 환경/제한사항을 명시합니다)"
           ).font = Font(italic=True, color="64748B", size=9)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    row += 2

    def _section(title: str) -> None:
        nonlocal row
        c = ws.cell(row=row, column=1, value=title)
        c.font = HDR_FONT
        c.fill = HDR_FILL
        c.alignment = Alignment(vertical="center", indent=1)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        ws.row_dimensions[row].height = 22
        row += 1

    def _kv(key: str, val) -> None:
        nonlocal row
        k = ws.cell(row=row, column=1, value=key)
        k.font = KEY_FONT
        k.alignment = Alignment(vertical="center", indent=1)
        k.border = BORDER
        v = ws.cell(row=row, column=2, value=_san(str(val)) if val not in (None, "") else "—")
        v.alignment = Alignment(vertical="center", wrap_text=True)
        v.border = BORDER
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
        row += 1

    # ── 1. 시험 환경 ───────────────────────────────────────────────────
    _section("1. 시험 환경 (재현 가능성 정보)")
    _kv("Run ID",               meta.get("run_id"))
    _kv("시작 일시",             meta.get("created_at"))
    _kv("종료 일시",             meta.get("updated_at"))
    _kv("대상 URL",              meta.get("target_url"))
    _kv("최종 단계",             meta.get("stage"))
    _kv("LLM 모델 (기본)",       meta.get("model_override") or "(contract 기본값)")
    _mo = meta.get("model_overrides") or {}
    if _mo:
        _kv("단계별 모델",
            ", ".join(f"{k}={v}" for k, v in _mo.items()))
    _kv("브라우저 모드",         "헤드리스" if meta.get("headless_exec", True) else "헤드풀(표시)")
    _kv("슬로우 모드(ms)",       meta.get("slow_mo_ms", 0))
    _kv("INFERRED 임계값",       meta.get("inferred_threshold"))
    _kv("max_leaves 설정",       meta.get("max_leaves"))
    _kv("max_pages 설정",        meta.get("max_pages"))
    _kv("입력 파일",             ", ".join(meta.get("input_files") or []) or "(없음)")
    row += 1

    # ── 2. 분석 범위 ───────────────────────────────────────────────────
    _section("2. 분석 범위 (선택된 페이지 / 캐시 재사용 / 중복 정리)")
    sel_urls = meta.get("selected_urls") or []
    cache_urls = meta.get("dom_cache_used") or []
    url_groups = meta.get("selected_url_groups") or {}
    n_merged = sum(len(v) for v in url_groups.values())
    _kv("분석 대상 페이지 수",   f"{len(sel_urls)}개" if sel_urls else "(전체 BFS)")
    _kv("캐시 재사용 페이지 수", f"{len(cache_urls)}개")
    _kv("중복 정리(동형 묶음)",
        f"{len(url_groups)}개 대표가 동형 {n_merged}개 포함" if n_merged else "없음")

    # 동형 묶음 상세 — 대표 1개로 동형 N개를 대표 시험했음을 명시 (커버리지 정당화)
    if url_groups:
        sub = ws.cell(row=row, column=1,
                      value="  동형 페이지 묶음 (대표 1개 시험 = 동형 전체 적용):")
        sub.font = SUB_FONT
        sub.fill = SUB_FILL
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        row += 1
        for rep, members in list(url_groups.items())[:30]:
            c1 = ws.cell(row=row, column=1, value=f"  대표(+{len(members)})")
            c1.font = Font(size=9, color="1E6B3C", bold=True)
            c2 = ws.cell(row=row, column=2, value=_san(rep))
            c2.font = Font(size=9, color="475569")
            c2.alignment = Alignment(wrap_text=True)
            c3 = ws.cell(row=row, column=3,
                         value=_san(f"동형 {len(members)}개: " + ", ".join(members[:3])
                               + (" …" if len(members) > 3 else "")))
            c3.font = Font(size=8, color="94A3B8")
            c3.alignment = Alignment(wrap_text=True)
            row += 1
        if len(url_groups) > 30:
            c = ws.cell(row=row, column=1, value=f"  … 외 {len(url_groups) - 30}개 그룹")
            c.font = Font(italic=True, color="64748B", size=9)
            row += 1
    row += 1

    if cache_urls:
        sub = ws.cell(row=row, column=1, value="  캐시 재사용 URL 목록:")
        sub.font = SUB_FONT
        sub.fill = SUB_FILL
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        row += 1
        for u in cache_urls:
            c = ws.cell(row=row, column=2, value=_san(u))
            c.alignment = Alignment(wrap_text=True)
            c.font = Font(size=9, color="475569")
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
            row += 1
    row += 1

    # ── 3. 시험 커버리지 ───────────────────────────────────────────────
    _section("3. 시험 커버리지 (기능 정제 → TC 설계 비율)")
    refine = meta.get("refine_report") or {}
    consol = meta.get("consolidate_report") or {}
    cov = meta.get("coverage") or {}
    if refine:
        _kv("Stage 0 원본 기능 수",  f"{refine.get('original','?')}개")
        _kv("노이즈 제외(UI 동작)",  f"{refine.get('removed_noise',0)}개")
        _kv("규칙 중복 병합",        f"{refine.get('merged_dup',0)}개")
        _kv("규칙 정제 후",          f"{refine.get('final','?')}개")
    if consol and not consol.get("skipped"):
        _kv("LLM 의미 통합 전",      f"{consol.get('before','?')}개")
        _kv("LLM 의미 통합 후(고유)", f"{consol.get('after','?')}개")
    if cov:
        _kv("TC 설계된 고유 기능",  f"{cov.get('designed_features','?')}개")
        _kv("총 TC 수",            f"{cov.get('total_tcs','?')}개")
        pct = cov.get("coverage_pct", 0)
        cov_cell_row = row
        _kv("✅ 시험 커버리지",     f"{pct}%  (고유 기능 대비 TC 설계율)")
        # 커버리지 색상 강조
        try:
            c = ws.cell(row=cov_cell_row, column=2)
            if pct >= 90:
                c.fill = PatternFill("solid", fgColor="C6EFCE")   # 연두
            elif pct >= 50:
                c.fill = PatternFill("solid", fgColor="FFEB9C")   # 노랑
            else:
                c.fill = PatternFill("solid", fgColor="FFC7CE")   # 빨강
            c.font = Font(bold=True, size=10)
        except Exception:
            pass
    row += 1

    # ── 4. 분석 누락 (실패 + 제외) ─────────────────────────────────────
    _section("4. 분석에서 제외/실패한 항목 (커버리지 미달 사유)")
    failed   = meta.get("stage2_failed_leaves") or []
    excluded = meta.get("stage2_excluded_leaves") or []
    _kv("분석 실패 leaf 수",  f"{len(failed)}개")
    _kv("max_leaves cap으로 제외", f"{len(excluded)}개")

    if failed:
        sub = ws.cell(row=row, column=1,
                      value="  분석 실패 leaf (LLM 안전 필터 / 응답 오류 등):")
        sub.font = SUB_FONT
        sub.fill = SUB_FILL
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        row += 1
        # 헤더
        for col_i, h in enumerate(["#", "leaf 명칭", "사유"], 1):
            c = ws.cell(row=row, column=col_i, value=h)
            c.font = Font(bold=True, color="475569", size=9)
            c.fill = SUB_FILL
            c.border = BORDER
        row += 1
        for item in failed[:50]:    # 최대 50개
            for col_i, val in enumerate([
                item.get("idx"), item.get("name"), item.get("reason"),
            ], 1):
                c = ws.cell(row=row, column=col_i, value=_san(val))
                c.font = Font(size=9)
                c.border = BORDER
                c.alignment = Alignment(wrap_text=True, vertical="top")
            row += 1
        if len(failed) > 50:
            c = ws.cell(row=row, column=1, value=f"  … 외 {len(failed) - 50}건 더 있음")
            c.font = Font(italic=True, color="64748B", size=9)
            row += 1
        row += 1

    if excluded:
        sub = ws.cell(row=row, column=1,
                      value="  max_leaves 우선순위 컷오프로 제외된 leaf:")
        sub.font = SUB_FONT
        sub.fill = SUB_FILL
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        row += 1
        for col_i, h in enumerate(["#", "leaf 명칭", "신뢰도"], 1):
            c = ws.cell(row=row, column=col_i, value=h)
            c.font = Font(bold=True, color="475569", size=9)
            c.fill = SUB_FILL
            c.border = BORDER
        row += 1
        for item in excluded[:50]:
            for col_i, val in enumerate([
                item.get("idx"), item.get("name"), item.get("confidence"),
            ], 1):
                c = ws.cell(row=row, column=col_i, value=_san(val))
                c.font = Font(size=9)
                c.border = BORDER
            row += 1
        if len(excluded) > 50:
            c = ws.cell(row=row, column=1, value=f"  … 외 {len(excluded) - 50}건 더 있음")
            c.font = Font(italic=True, color="64748B", size=9)
            row += 1
    row += 1

    # ── 4. 결과 요약 ───────────────────────────────────────────────────
    _section("4. 결과 요약")
    n_total = len(tcs)
    n_pass = sum(1 for t in tcs if t.get("result") == "pass")
    n_fail = sum(1 for t in tcs if t.get("result") == "fail")
    n_blocked = sum(1 for t in tcs if t.get("result") == "blocked")
    n_not_run = sum(1 for t in tcs if t.get("result") in ("not_executed", "", None))
    _kv("총 TC 수",      f"{n_total}개")
    _kv("통과 (PASS)",   f"{n_pass}개")
    _kv("실패 (FAIL)",   f"{n_fail}개")
    _kv("차단 (BLOCKED)", f"{n_blocked}개")
    _kv("미실행",         f"{n_not_run}개")

    # 열 너비
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 52
    ws.column_dimensions["C"].width = 40


def build_review(tcs: list[dict], output_path: str | Path) -> Path:
    """Reviewer Gate용 xlsx (Stage 4 전 단계)."""
    review_cols = (
        _SHEET1_COLS[:7]
        + ["design_technique", "source_quote", "gen_confidence",
           "review_status", "reviewer_note", "screenshot_file"]
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "TC 검토"
    _write_sheet(ws, review_cols, tcs, confidence_col="gen_confidence")

    # screenshot_file 열 너비
    if "screenshot_file" in review_cols:
        sf_col = get_column_letter(review_cols.index("screenshot_file") + 1)
        ws.column_dimensions[sf_col].width = 30

    # 드롭다운 (review_status 컬럼) — TC가 1개 이상일 때만 설정
    # len(tcs)==0 이면 sqref="K2:K1" → min_row>max_row → openpyxl ValueError
    if tcs:
        from openpyxl.worksheet.datavalidation import DataValidation
        status_col_idx    = review_cols.index("review_status") + 1
        status_col_letter = get_column_letter(status_col_idx)
        dv = DataValidation(
            type="list",
            formula1='"approved,edited,rejected,pending"',
            allow_blank=False,
        )
        ws.add_data_validation(dv)
        dv.sqref = f"{status_col_letter}2:{status_col_letter}{len(tcs) + 1}"

    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out))
    return out


# ── 기능 목록 Excel (Stage 0 DOM 스캔 결과) ───────────────────────────────────

def build_features(features: list[dict], output_path: str | Path) -> Path:
    """Stage 0 기능 목록을 Excel로 저장.

    Args:
        features: feature-spec-draft.json 의 'features' 리스트
        output_path: 저장 경로 (.xlsx)
    Returns:
        저장된 파일 경로
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "기능 목록"
    ws.freeze_panes = "A2"

    # 헤더
    for ci, col in enumerate(_FEATURE_COLS, 1):
        cell = ws.cell(row=1, column=ci, value=_FEATURE_COL_NAMES.get(col, col))
        _header_style(cell, fill_color="1F4E79")  # 진한 남색 — TC Excel과 구분
        ws.column_dimensions[get_column_letter(ci)].width = _FEATURE_COL_WIDTHS.get(col, 16)

    # 같은 분류끼리 인접하도록 정렬 (병합셀 전제) — 대분류 > 중분류 > 소분류
    feats_sorted = sorted(
        features,
        key=lambda f: (
            str(f.get("category_major", "")),
            str(f.get("category_mid", "")),
            str(f.get("category_leaf", "")),
        ),
    )

    # 데이터
    for ri, feat in enumerate(feats_sorted, 2):
        for ci, col in enumerate(_FEATURE_COLS, 1):
            val  = feat.get(col, "")
            cell = ws.cell(row=ri, column=ci, value=_san(val))
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border    = _BORDER

            # confidence 색상 (HIGH/MID/INFERRED)
            if col == "confidence" and isinstance(val, str):
                val_upper = val.upper()
                if val_upper == "HIGH":
                    cell.fill = _CONFIDENCE_FILLS["high"]
                elif val_upper == "MID":
                    cell.fill = _CONFIDENCE_FILLS["mid"]
                elif val_upper == "INFERRED":
                    cell.fill = _CONFIDENCE_FILLS["low"]

    # ── 같은 분류 세로 병합셀 (대분류=1열, 중분류=2열) — 그룹 가독성 ──────────
    n = len(feats_sorted)
    if n >= 2:
        merged_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

        def _merge_runs(col_idx: int, keyfn) -> None:
            start = 2                       # 데이터 첫 행(헤더 다음)
            prev  = keyfn(feats_sorted[0])
            for i in range(1, n):
                cur = keyfn(feats_sorted[i])
                row = i + 2
                if cur != prev:
                    if row - 1 > start:     # run 길이 ≥ 2일 때만 병합
                        ws.merge_cells(start_row=start, start_column=col_idx,
                                       end_row=row - 1, end_column=col_idx)
                        ws.cell(row=start, column=col_idx).alignment = merged_align
                    start = row
                    prev  = cur
            last = n + 1
            if last > start:
                ws.merge_cells(start_row=start, start_column=col_idx,
                               end_row=last, end_column=col_idx)
                ws.cell(row=start, column=col_idx).alignment = merged_align

        # 1열: 대분류 / 2열: (대분류,중분류) 동일 구간
        _merge_runs(1, lambda f: str(f.get("category_major", "")))
        _merge_runs(2, lambda f: (str(f.get("category_major", "")),
                                  str(f.get("category_mid", ""))))

    out = Path(output_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out))
    return out
